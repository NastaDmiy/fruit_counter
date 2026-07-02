from flask import Flask, render_template, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import os
import json
import sqlite3
from datetime import datetime
from io import BytesIO
from detector import detect_fruits
from openpyxl import Workbook

app = Flask(__name__)
CORS(app)

# Создаём папки, если их нет
UPLOAD_FOLDER = 'uploads'
STATIC_FOLDER = 'static'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(STATIC_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# --- База данных SQLite ---
def init_db():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            filename TEXT,
            total_count INTEGER,
            details TEXT
        )
    ''')
    conn.commit()
    conn.close()

init_db()

def save_to_history(filename, total_count, details):
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute(
        'INSERT INTO history (timestamp, filename, total_count, details) VALUES (?, ?, ?, ?)',
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), filename, total_count, json.dumps(details, ensure_ascii=False))
    )
    conn.commit()
    conn.close()

# --- Маршруты ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_image():
    if 'image' not in request.files:
        return jsonify({'error': 'Файл не загружен'}), 400

    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'Имя файла пустое'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        total_count, fruit_details, output_path = detect_fruits(filepath)
        save_to_history(file.filename, total_count, fruit_details)

        return jsonify({
            'success': True,
            'total_count': total_count,
            'details': fruit_details,
            'result_image': output_path,
            'filename': file.filename
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/history')
def get_history():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT timestamp, filename, total_count, details FROM history ORDER BY id DESC LIMIT 20')
    rows = cursor.fetchall()
    conn.close()

    history = []
    for row in rows:
        history.append({
            'timestamp': row[0],
            'filename': row[1],
            'total_count': row[2],
            'details': json.loads(row[3]) if row[3] else {}
        })
    return jsonify(history)

@app.route('/export_excel')
def export_excel():
    """Экспорт истории в Excel"""
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT timestamp, filename, total_count, details FROM history ORDER BY id DESC')
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "История обработки"

    ws['A1'] = 'Дата и время'
    ws['B1'] = 'Имя файла'
    ws['C1'] = 'Всего фруктов'
    ws['D1'] = 'Детали'

    for i, row in enumerate(rows, start=2):
        ws[f'A{i}'] = row[0]
        ws[f'B{i}'] = row[1]
        ws[f'C{i}'] = row[2]
        ws[f'D{i}'] = row[3]

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='history_fruits.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(STATIC_FOLDER, filename)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)